from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
import io
import csv
from openpyxl import load_workbook
from app.database import get_db
from app.schemas.task import (
    TaskCreate, TaskUpdate, Task, TaskWithAssignments,
    TaskAssignmentCreate, TaskAssignmentUpdate, TaskAssignmentResponse,
    TaskBulkImportResult
)
from app.services.task_service import TaskService, TaskAssignmentService
from app.api.deps import get_current_user
from app.models.user import User

router = APIRouter()

# ======================== 任务管理接口 ========================

@router.get("/tasks", response_model=List[Task])
async def get_tasks(
    skip: int = Query(0, ge=0, description="跳过的记录数"),
    limit: int = Query(100, ge=1, le=1000, description="限制返回的记录数"),
    status: Optional[str] = Query(None, description="任务状态过滤"),
    task_type: Optional[str] = Query(None, description="任务类型过滤"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取任务列表"""
    tasks = TaskService.get_tasks(
        db=db, 
        skip=skip, 
        limit=limit, 
        status=status, 
        task_type=task_type
    )
    return tasks

@router.get("/tasks/count")
async def get_tasks_count(
    status: Optional[str] = Query(None, description="任务状态过滤"),
    task_type: Optional[str] = Query(None, description="任务类型过滤"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取任务总数"""
    count = TaskService.get_tasks_count(db=db, status=status, task_type=task_type)
    return {"count": count}

@router.get("/tasks/{task_id}", response_model=TaskWithAssignments)
async def get_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """根据ID获取任务详情（包含分配信息）"""
    task = TaskService.get_task_with_assignments(db=db, task_id=task_id)
    if not task:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="任务不存在"
        )
    return task

@router.post("/tasks", response_model=Task)
async def create_task(
    task: TaskCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建任务"""
    # 检查权限（只有管理员可以创建任务）
    if current_user.type != "管理员":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以创建任务"
        )
    
    db_task = TaskService.create_task(db=db, task=task)
    return db_task

@router.put("/tasks/{task_id}", response_model=Task)
async def update_task(
    task_id: int,
    task: TaskUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新任务"""
    # 检查权限（只有管理员可以更新任务）
    if current_user.type != "管理员":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以更新任务"
        )
    
    db_task = TaskService.update_task(db=db, task_id=task_id, task=task)
    if not db_task:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="任务不存在"
        )
    return db_task

@router.delete("/tasks/{task_id}")
async def delete_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除任务"""
    # 检查权限（只有管理员可以删除任务）
    if current_user.type != "管理员":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以删除任务"
        )
    
    success = TaskService.delete_task(db=db, task_id=task_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="任务不存在"
        )
    return {"message": "任务删除成功"}

@router.post("/tasks/bulk-import", response_model=TaskBulkImportResult)
async def bulk_import_tasks(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> TaskBulkImportResult:
    """批量导入任务（管理员权限）"""
    # 检查权限（只有管理员可以批量导入任务）
    if current_user.type != "管理员":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以批量导入任务"
        )
    
    # 验证文件名
    if not file.filename:
        raise HTTPException(
            status_code=400,
            detail="文件名不能为空"
        )
    
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=400,
            detail="只支持Excel文件格式（.xlsx, .xls）或CSV文件格式（.csv）"
        )
    
    # 验证文件大小
    if file.size and file.size > 10 * 1024 * 1024:  # 10MB
        raise HTTPException(
            status_code=400,
            detail="文件大小不能超过10MB"
        )
    
    try:
        # 读取文件内容
        contents = await file.read()
        
        # 验证文件内容不为空
        if not contents:
            raise HTTPException(
                status_code=400,
                detail="文件内容为空"
            )
        
        # 判断文件类型并处理
        is_csv = file.filename.endswith('.csv')
        
        if is_csv:
            # 处理CSV文件
            try:
                content_str = contents.decode('utf-8-sig')  # 处理BOM
            except UnicodeDecodeError:
                try:
                    content_str = contents.decode('gbk')  # 尝试GBK编码
                except UnicodeDecodeError:
                    content_str = contents.decode('utf-8', errors='ignore')
            
            # 使用CSV reader解析
            csv_reader = csv.reader(io.StringIO(content_str))
            rows = list(csv_reader)
            
            if len(rows) < 2:
                raise HTTPException(
                    status_code=400,
                    detail="CSV文件为空或只有表头，请添加任务数据"
                )
            
            # 获取表头
            headers = [col.strip() for col in rows[0] if col.strip()]
            
            if not headers:
                raise HTTPException(
                    status_code=400,
                    detail="CSV文件表头为空"
                )
            
            # 验证必要的列
            required_columns = ['任务名称', '任务类型', '阶段', '任务描述']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"CSV文件缺少必要列：{', '.join(missing_columns)}。请使用正确的模板格式。"
                )
            
            # 获取列索引
            column_indexes = {col: headers.index(col) for col in required_columns}
            
            # 转换为TaskCreate对象列表
            tasks_to_create = []
            invalid_rows = []
            
            for row_num, row in enumerate(rows[1:], start=2):
                if not row or not any(row):  # 跳过空行
                    continue
                    
                try:
                    # 获取各字段值
                    task_name = row[column_indexes['任务名称']].strip() if len(row) > column_indexes['任务名称'] and row[column_indexes['任务名称']] else ""
                    task_type = row[column_indexes['任务类型']].strip() if len(row) > column_indexes['任务类型'] and row[column_indexes['任务类型']] else ""
                    task_phase = row[column_indexes['阶段']].strip() if len(row) > column_indexes['阶段'] and row[column_indexes['阶段']] else ""
                    task_description = row[column_indexes['任务描述']].strip() if len(row) > column_indexes['任务描述'] and row[column_indexes['任务描述']] else ""
                    
                    # 验证必填字段
                    if not task_name:
                        invalid_rows.append(f"第{row_num}行：任务名称不能为空")
                        continue
                    if not task_type:
                        invalid_rows.append(f"第{row_num}行：任务类型不能为空")
                        continue
                    if not task_phase:
                        invalid_rows.append(f"第{row_num}行：阶段不能为空")
                        continue
                    if not task_description:
                        invalid_rows.append(f"第{row_num}行：任务描述不能为空")
                        continue
                    
                    task_create = TaskCreate(
                        name=task_name,
                        type=task_type,
                        phase=task_phase,
                        description=task_description,
                        status='未分配'
                    )
                    tasks_to_create.append(task_create)
                    
                except (IndexError, ValueError, TypeError) as row_error:
                    invalid_rows.append(f"第{row_num}行：数据格式错误 - {str(row_error)}")
                    continue
        
        else:
            # 处理Excel文件
            try:
                workbook = load_workbook(io.BytesIO(contents))
            except Exception as excel_error:
                # 提供更详细的Excel文件解析错误信息
                error_msg = str(excel_error)
                if "not a zip file" in error_msg.lower():
                    raise HTTPException(
                        status_code=400,
                        detail="文件格式错误：请确保上传的是有效的Excel文件（.xlsx或.xls）"
                    )
                elif "corrupted" in error_msg.lower() or "damaged" in error_msg.lower():
                    raise HTTPException(
                        status_code=400,
                        detail="文件已损坏：请重新下载模板并填写数据"
                    )
                else:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Excel文件解析失败：{error_msg}"
                    )
            
            worksheet = workbook.active
            
            # 检查工作表是否为空
            if worksheet.max_row < 2:
                raise HTTPException(
                    status_code=400,
                    detail="Excel文件为空或只有表头，请添加任务数据"
                )
            
            # 获取表头
            headers = [cell.value for cell in worksheet[1] if cell.value]
            
            if not headers:
                raise HTTPException(
                    status_code=400,
                    detail="Excel文件表头为空"
                )
            
            # 验证必要的列
            required_columns = ['任务名称', '任务类型', '阶段', '任务描述']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"Excel文件缺少必要列：{', '.join(missing_columns)}。请使用正确的模板格式。"
                )
            
            # 获取列索引
            column_indexes = {col: headers.index(col) for col in required_columns}
            
            # 转换为TaskCreate对象列表
            tasks_to_create = []
            invalid_rows = []
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):  # 跳过空行
                    continue
                    
                try:
                    # 获取各字段值
                    task_name = str(row[column_indexes['任务名称']]).strip() if row[column_indexes['任务名称']] else ""
                    task_type = str(row[column_indexes['任务类型']]).strip() if row[column_indexes['任务类型']] else ""
                    task_phase = str(row[column_indexes['阶段']]).strip() if row[column_indexes['阶段']] else ""
                    task_description = str(row[column_indexes['任务描述']]).strip() if row[column_indexes['任务描述']] else ""
                    
                    # 验证必填字段
                    if not task_name:
                        invalid_rows.append(f"第{row_num}行：任务名称不能为空")
                        continue
                    if not task_type:
                        invalid_rows.append(f"第{row_num}行：任务类型不能为空")
                        continue
                    if not task_phase:
                        invalid_rows.append(f"第{row_num}行：阶段不能为空")
                        continue
                    if not task_description:
                        invalid_rows.append(f"第{row_num}行：任务描述不能为空")
                        continue
                    
                    task_create = TaskCreate(
                        name=task_name,
                        type=task_type,
                        phase=task_phase,
                        description=task_description,
                        status='未分配'
                    )
                    tasks_to_create.append(task_create)
                    
                except (IndexError, ValueError, TypeError) as row_error:
                    invalid_rows.append(f"第{row_num}行：数据格式错误 - {str(row_error)}")
                    continue
        
        # 如果有无效行，提供详细信息
        if invalid_rows and not tasks_to_create:
            raise HTTPException(
                status_code=400,
                detail=f"所有数据行都无效：{'; '.join(invalid_rows[:5])}" + ("..." if len(invalid_rows) > 5 else "")
            )
        
        if not tasks_to_create:
            raise HTTPException(
                status_code=400,
                detail="文件中没有有效的任务数据"
            )
        
        # 批量创建任务
        result = TaskService.bulk_create_tasks(db, tasks_to_create)
        
        # 如果有无效行，在结果中提醒
        if invalid_rows:
            result.message += f"（注意：跳过了{len(invalid_rows)}行无效数据）"
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        # 捕获所有其他异常并提供详细错误信息
        error_msg = str(e)
        if "openpyxl" in error_msg.lower():
            raise HTTPException(
                status_code=400,
                detail="Excel文件处理库出错，请检查文件格式是否正确"
            )
        else:
            raise HTTPException(
                status_code=400,
                detail=f"文件处理失败：{error_msg}"
            )

# ======================== 任务分配接口 ========================

@router.get("/task-assignments", response_model=List[TaskAssignmentResponse])
async def get_task_assignments(
    skip: int = Query(0, ge=0, description="跳过的记录数"),
    limit: int = Query(100, ge=1, le=1000, description="限制返回的记录数"),
    status: Optional[str] = Query(None, description="分配状态过滤"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取任务分配列表"""
    assignments = TaskAssignmentService.get_assignments(
        db=db, 
        skip=skip, 
        limit=limit, 
        status=status
    )
    return assignments

@router.get("/tasks/{task_id}/assignments", response_model=List[TaskAssignmentResponse])
async def get_task_assignments_by_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """根据任务ID获取分配列表"""
    # 检查任务是否存在
    task = TaskService.get_task(db=db, task_id=task_id)
    if not task:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="任务不存在"
        )
    
    assignments = TaskAssignmentService.get_assignments_by_task(db=db, task_id=task_id)
    return assignments

@router.get("/users/{user_id}/task-assignments", response_model=List[TaskAssignmentResponse])
async def get_user_task_assignments(
    user_id: int,
    status: Optional[str] = Query(None, description="分配状态过滤"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """根据用户ID获取任务分配列表"""
    # 检查权限（用户只能查看自己的分配，管理员可以查看所有）
    if current_user.type != "管理员" and current_user.id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权查看其他用户的任务分配"
        )
    
    assignments = TaskAssignmentService.get_assignments_by_user(
        db=db, 
        user_id=user_id, 
        status=status
    )
    return assignments

@router.get("/users/{user_id}/task-stats")
async def get_user_task_stats(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取用户任务统计"""
    # 检查权限（用户只能查看自己的统计，管理员可以查看所有）
    if current_user.type != "管理员" and current_user.id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权查看其他用户的任务统计"
        )
    
    stats = TaskAssignmentService.get_user_task_stats(db=db, user_id=user_id)
    return stats

@router.get("/task-assignments/stats")
async def get_task_assignments_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取全部任务分配统计"""
    # 检查权限（只有管理员可以查看全部统计）
    if current_user.type != "管理员":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以查看全部任务统计"
        )
    
    stats = TaskAssignmentService.get_all_task_stats(db=db)
    return stats

@router.post("/task-assignments", response_model=TaskAssignmentResponse)
async def create_task_assignment(
    assignment: TaskAssignmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建任务分配"""
    # 检查权限（只有管理员可以分配任务）
    if current_user.type != "管理员":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以分配任务"
        )
    
    db_assignment = TaskAssignmentService.create_assignment(db=db, assignment=assignment)
    if not db_assignment:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="任务或用户不存在"
        )
    return db_assignment

@router.put("/task-assignments/{assignment_id}", response_model=TaskAssignmentResponse)
async def update_task_assignment(
    assignment_id: int,
    assignment: TaskAssignmentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新任务分配"""
    # 获取现有分配
    existing_assignment = TaskAssignmentService.get_assignment(db=db, assignment_id=assignment_id)
    if not existing_assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="任务分配不存在"
        )
    
    # 检查权限（用户只能更新自己的分配，管理员可以更新所有）
    if current_user.type != "管理员" and current_user.id != existing_assignment.user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权更新其他用户的任务分配"
        )
    
    db_assignment = TaskAssignmentService.update_assignment(
        db=db, 
        assignment_id=assignment_id, 
        assignment=assignment
    )
    return db_assignment

@router.delete("/task-assignments/{assignment_id}")
async def delete_task_assignment(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除任务分配"""
    # 检查权限（只有管理员可以删除任务分配）
    if current_user.type != "管理员":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有管理员可以删除任务分配"
        )
    
    success = TaskAssignmentService.delete_assignment(db=db, assignment_id=assignment_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="任务分配不存在"
        )
    return {"message": "任务分配删除成功"}

# ======================== 我的任务接口 ========================

@router.get("/my-tasks", response_model=List[TaskAssignmentResponse])
async def get_my_tasks(
    status: Optional[str] = Query(None, description="任务状态过滤"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取当前用户的任务分配"""
    assignments = TaskAssignmentService.get_assignments_by_user(
        db=db, 
        user_id=current_user.id, 
        status=status
    )
    return assignments

@router.get("/my-task-stats")
async def get_my_task_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取当前用户的任务统计"""
    stats = TaskAssignmentService.get_user_task_stats(db=db, user_id=current_user.id)
    return stats

@router.put("/my-tasks/{assignment_id}", response_model=TaskAssignmentResponse)
async def update_my_task(
    assignment_id: int,
    assignment: TaskAssignmentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新我的任务进度"""
    # 获取现有分配
    existing_assignment = TaskAssignmentService.get_assignment(db=db, assignment_id=assignment_id)
    if not existing_assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="任务分配不存在"
        )
    
    # 检查权限（只能更新自己的任务）
    if current_user.id != existing_assignment.user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权更新其他用户的任务"
        )
    
    db_assignment = TaskAssignmentService.update_assignment(
        db=db, 
        assignment_id=assignment_id, 
        assignment=assignment
    )
    return db_assignment 