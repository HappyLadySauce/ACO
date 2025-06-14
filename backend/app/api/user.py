from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
import io
import json
import csv
from openpyxl import Workbook, load_workbook
from app.database import get_db
from app.schemas.user import UserCreate, UserResponse, UserUpdate, BulkImportResult
from app.services.user_service import UserService
from app.api.deps import get_current_user, get_admin_user

router = APIRouter()

@router.get("/users", response_model=List[UserResponse])
async def get_users(
    skip: int = Query(0, ge=0, description="跳过的记录数"),
    limit: int = Query(100, ge=1, le=1000, description="返回的记录数"),
    role: str = Query(None, description="筛选角色"),
    user_type: str = Query(None, description="筛选用户类型"),
    db: Session = Depends(get_db),
    current_user = Depends(get_admin_user)
) -> List[UserResponse]:
    """获取用户列表（管理员权限）"""
    if role:
        users = UserService.get_users_by_role(db, role)
    elif user_type:
        users = UserService.get_users_by_type(db, user_type)
    else:
        users = UserService.get_users(db, skip=skip, limit=limit)
    
    return [UserResponse.model_validate(user) for user in users]

@router.post("/users", response_model=UserResponse)
async def create_user(
    user: UserCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_admin_user)
) -> UserResponse:
    """创建用户（管理员权限）"""
    db_user = UserService.get_user_by_username(db, username=user.username)
    if db_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="用户名已存在"
        )
    
    new_user = UserService.create_user(db=db, user=user)
    return UserResponse.model_validate(new_user)

@router.get("/users/{user_id}", response_model=UserResponse)
async def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_admin_user)
) -> UserResponse:
    """获取指定用户信息（管理员权限）"""
    db_user = UserService.get_user(db, user_id=user_id)
    if not db_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )
    return UserResponse.model_validate(db_user)

@router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    user: UserUpdate,
    db: Session = Depends(get_db),
    current_user = Depends(get_admin_user)
) -> UserResponse:
    """更新用户信息（管理员权限）"""
    db_user = UserService.get_user(db, user_id=user_id)
    if not db_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )
    
    updated_user = UserService.update_user(db=db, user_id=user_id, user=user)
    return UserResponse.model_validate(updated_user)

@router.delete("/users/{user_id}")
async def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_admin_user)
):
    """删除用户（管理员权限）"""
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="不能删除自己的账户"
        )
    
    success = UserService.delete_user(db=db, user_id=user_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )
    
    return {"message": "用户删除成功"}

@router.patch("/users/{user_id}/status")
async def update_user_status(
    user_id: int,
    status: str = Query(..., description="用户状态"),
    db: Session = Depends(get_db),
    current_user = Depends(get_admin_user)
) -> UserResponse:
    """更新用户状态（管理员权限）"""
    if user_id == current_user.id and status == 'inactive':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="不能禁用自己的账户"
        )
    
    updated_user = UserService.update_user_status(db=db, user_id=user_id, status=status)
    if not updated_user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="用户不存在"
        )
    
    return UserResponse.model_validate(updated_user)

@router.get("/me", response_model=UserResponse)
async def get_current_user_info(current_user = Depends(get_current_user)) -> UserResponse:
    """获取当前用户信息"""
    return UserResponse.model_validate(current_user)

@router.put("/profile", response_model=UserResponse)
async def update_profile(
    user: UserUpdate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
) -> UserResponse:
    """更新个人资料"""
    # 普通用户只能更新自己的头像等基本信息
    profile_data = UserUpdate(photo_data=user.photo_data)
    updated_user = UserService.update_user(db=db, user_id=current_user.id, user=profile_data)
    return UserResponse.model_validate(updated_user)

@router.post("/users/bulk-import", response_model=BulkImportResult)
async def bulk_import_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_admin_user)
) -> BulkImportResult:
    """批量导入用户（管理员权限）"""
    # 验证文件名
    if not file.filename:
        raise HTTPException(
            status_code=400,
            detail="文件名不能为空"
        )
    
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(
            status_code=400,
            detail="只支持Excel文件格式（.xlsx, .xls）或CSV文件格式（.csv）"
        )
    
    # 验证文件大小
    if file.size and file.size > 10 * 1024 * 1024:  # 10MB
        raise HTTPException(
            status_code=400,
            detail="文件大小不能超过10MB"
        )
    
    try:
        # 读取Excel文件
        contents = await file.read()
        
        # 验证文件内容不为空
        if not contents:
            raise HTTPException(
                status_code=400,
                detail="文件内容为空"
            )
        
        # 判断文件类型并处理
        is_csv = file.filename.endswith('.csv')
        
        if is_csv:
            # 处理CSV文件
            try:
                content_str = contents.decode('utf-8-sig')  # 处理BOM
            except UnicodeDecodeError:
                try:
                    content_str = contents.decode('gbk')  # 尝试GBK编码
                except UnicodeDecodeError:
                    content_str = contents.decode('utf-8', errors='ignore')
            
            # 使用CSV reader解析
            csv_reader = csv.reader(io.StringIO(content_str))
            rows = list(csv_reader)
            
            if len(rows) < 2:
                raise HTTPException(
                    status_code=400,
                    detail="CSV文件为空或只有表头，请添加用户数据"
                )
            
            # 获取表头
            headers = [col.strip() for col in rows[0] if col.strip()]
            
            if not headers:
                raise HTTPException(
                    status_code=400,
                    detail="CSV文件表头为空"
                )
            
            # 验证必要的列
            required_columns = ['用户名', '密码', '角色', '用户类型', '状态']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"CSV文件缺少必要列：{', '.join(missing_columns)}。请使用正确的模板格式。"
                )
            
            # 获取列索引
            column_indexes = {col: headers.index(col) for col in required_columns}
            
            # 转换为UserCreate对象列表
            users_to_create = []
            invalid_rows = []
            
            for row_num, row in enumerate(rows[1:], start=2):
                if not row or not any(row):  # 跳过空行
                    continue
                    
                try:
                    # 获取各字段值
                    username = row[column_indexes['用户名']].strip() if len(row) > column_indexes['用户名'] and row[column_indexes['用户名']] else ""
                    password = row[column_indexes['密码']].strip() if len(row) > column_indexes['密码'] and row[column_indexes['密码']] else ""
                    role = row[column_indexes['角色']].strip() if len(row) > column_indexes['角色'] and row[column_indexes['角色']] else ""
                    user_type = row[column_indexes['用户类型']].strip() if len(row) > column_indexes['用户类型'] and row[column_indexes['用户类型']] else ""
                    user_status = row[column_indexes['状态']].strip() if len(row) > column_indexes['状态'] and row[column_indexes['状态']] else ""
                    
                    # 验证必填字段
                    if not username:
                        invalid_rows.append(f"第{row_num}行：用户名不能为空")
                        continue
                    if not password:
                        invalid_rows.append(f"第{row_num}行：密码不能为空")
                        continue
                    if not role:
                        invalid_rows.append(f"第{row_num}行：角色不能为空")
                        continue
                    if not user_type:
                        invalid_rows.append(f"第{row_num}行：用户类型不能为空")
                        continue
                    if not user_status:
                        invalid_rows.append(f"第{row_num}行：状态不能为空")
                        continue
                    
                    user_create = UserCreate(
                        username=username,
                        password=password,
                        role=role,
                        type=user_type,
                        status=user_status
                    )
                    users_to_create.append(user_create)
                    
                except (IndexError, ValueError, TypeError) as row_error:
                    invalid_rows.append(f"第{row_num}行：数据格式错误 - {str(row_error)}")
                    continue
        
        else:
            # 处理Excel文件
            try:
                workbook = load_workbook(io.BytesIO(contents))
            except Exception as excel_error:
                # 提供更详细的Excel文件解析错误信息
                error_msg = str(excel_error)
                if "not a zip file" in error_msg.lower():
                    raise HTTPException(
                        status_code=400,
                        detail="文件格式错误：请确保上传的是有效的Excel文件（.xlsx或.xls）"
                    )
                elif "corrupted" in error_msg.lower() or "damaged" in error_msg.lower():
                    raise HTTPException(
                        status_code=400,
                        detail="文件已损坏：请重新下载模板并填写数据"
                    )
                else:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Excel文件解析失败：{error_msg}"
                    )
            
            worksheet = workbook.active
            
            # 检查工作表是否为空
            if worksheet.max_row < 2:
                raise HTTPException(
                    status_code=400,
                    detail="Excel文件为空或只有表头，请添加用户数据"
                )
            
            # 获取表头
            headers = [cell.value for cell in worksheet[1] if cell.value]
            
            if not headers:
                raise HTTPException(
                    status_code=400,
                    detail="Excel文件表头为空"
                )
            
            # 验证必要的列
            required_columns = ['用户名', '密码', '角色', '用户类型', '状态']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"Excel文件缺少必要列：{', '.join(missing_columns)}。请使用正确的模板格式。"
                )
            
            # 获取列索引
            column_indexes = {col: headers.index(col) for col in required_columns}
            
            # 转换为UserCreate对象列表
            users_to_create = []
            invalid_rows = []
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):  # 跳过空行
                    continue
                    
                try:
                    # 获取各字段值
                    username = str(row[column_indexes['用户名']]).strip() if row[column_indexes['用户名']] else ""
                    password = str(row[column_indexes['密码']]).strip() if row[column_indexes['密码']] else ""
                    role = str(row[column_indexes['角色']]).strip() if row[column_indexes['角色']] else ""
                    user_type = str(row[column_indexes['用户类型']]).strip() if row[column_indexes['用户类型']] else ""
                    user_status = str(row[column_indexes['状态']]).strip() if row[column_indexes['状态']] else ""
                    
                    # 验证必填字段
                    if not username:
                        invalid_rows.append(f"第{row_num}行：用户名不能为空")
                        continue
                    if not password:
                        invalid_rows.append(f"第{row_num}行：密码不能为空")
                        continue
                    if not role:
                        invalid_rows.append(f"第{row_num}行：角色不能为空")
                        continue
                    if not user_type:
                        invalid_rows.append(f"第{row_num}行：用户类型不能为空")
                        continue
                    if not user_status:
                        invalid_rows.append(f"第{row_num}行：状态不能为空")
                        continue
                    
                    user_create = UserCreate(
                        username=username,
                        password=password,
                        role=role,
                        type=user_type,
                        status=user_status
                    )
                    users_to_create.append(user_create)
                    
                except (IndexError, ValueError, TypeError) as row_error:
                    invalid_rows.append(f"第{row_num}行：数据格式错误 - {str(row_error)}")
                    continue
        
        # 如果有无效行，提供详细信息
        if invalid_rows and not users_to_create:
            raise HTTPException(
                status_code=400,
                detail=f"所有数据行都无效：{'; '.join(invalid_rows[:5])}" + ("..." if len(invalid_rows) > 5 else "")
            )
        
        if not users_to_create:
            raise HTTPException(
                status_code=400,
                detail="文件中没有有效的用户数据"
            )
        
        # 批量创建用户
        result = UserService.bulk_create_users(db, users_to_create)
        
        # 如果有无效行，在结果中提醒
        if invalid_rows:
            result.message += f"（注意：跳过了{len(invalid_rows)}行无效数据）"
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        # 捕获所有其他异常并提供详细错误信息
        error_msg = str(e)
        if "openpyxl" in error_msg.lower():
            raise HTTPException(
                status_code=400,
                detail="Excel文件处理库出错，请检查文件格式是否正确"
            )
        else:
            raise HTTPException(
                status_code=400,
                detail=f"文件处理失败：{error_msg}"
            )

@router.get("/users/template/download")
async def download_user_template():
    """下载用户导入模板"""
    # 创建工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "用户导入模板"
    
    # 设置表头
    headers = ['用户名', '密码', '角色', '用户类型', '状态']
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # 添加示例数据
    sample_data = [
        ['user1', '123456', '网络工程师', '操作员', 'active'],
        ['user2', '123456', '系统架构师', '操作员', 'active'],
        ['user3', '123456', '系统规划与管理师', '管理员', 'active']
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存到内存
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=user_import_template.xlsx"}
    ) 